import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = ROOT / "Vocabulary.xlsx"
OUTPUT_PATH = ROOT / "public" / "data" / "vocabulary.json"

TARGET_SHEETS = OrderedDict(
    [
        ("War", "meaning"),
        ("Random", "meaning"),
        ("A World Undone", "meaning"),
        ("Competitive Exams", "bengali meaning"),
        ("Analogy", "bengali meaning"),
    ]
)


def clean_text(value: object, strip_terminal_punctuation: bool = False) -> str | None:
    if value is None or isinstance(value, bool):
        return None

    text = str(value).replace("\n", " ").replace("\u200d", "").strip()
    text = re.sub(r"\s+", " ", text)

    if strip_terminal_punctuation:
        text = re.sub(r"""[\s"'`.,;:!?।]+$""", "", text)

    return text or None


def build_dataset() -> dict:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    entries_by_key: OrderedDict[tuple[str, str], dict] = OrderedDict()
    source_counts: dict[str, int] = {}
    blank_rows: dict[str, int] = {}
    incomplete_rows: dict[str, int] = {}
    row_ranges: dict[str, dict[str, int]] = {}

    for sheet_name, meaning_header in TARGET_SHEETS.items():
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        header = rows[0]
        normalized_header = [
            clean_text(cell).casefold() if clean_text(cell) else "" for cell in header
        ]

        word_index = normalized_header.index("word")
        meaning_index = normalized_header.index(meaning_header)
        example_index = normalized_header.index("example") if "example" in normalized_header else None

        source_counts[sheet_name] = 0
        blank_rows[sheet_name] = 0
        incomplete_rows[sheet_name] = 0
        min_row = float('inf')
        max_row = 0

        for row_idx, row in enumerate(rows[1:], start=2):  # Start from row 2 (after header)
            values = list(row)
            visible_values = [
                clean_text(value) for value in values if clean_text(value) is not None
            ]
            if not visible_values:
                blank_rows[sheet_name] += 1
                continue

            word = clean_text(
                values[word_index] if word_index < len(values) else None,
                strip_terminal_punctuation=True,
            )
            bengali = clean_text(
                values[meaning_index] if meaning_index < len(values) else None,
                strip_terminal_punctuation=True,
            )
            example = clean_text(
                values[example_index] if example_index is not None and example_index < len(values) else None,
            )

            if not word or not bengali:
                incomplete_rows[sheet_name] += 1
                continue

            source_counts[sheet_name] += 1
            min_row = min(min_row, row_idx)
            max_row = max(max_row, row_idx)
            key = (word.casefold(), bengali.casefold())

            if key not in entries_by_key:
                entries_by_key[key] = {
                    "id": len(entries_by_key) + 1,
                    "english": word,
                    "bengali": bengali,
                    "example": example,
                    "sources": {sheet_name: row_idx},
                }
                continue

            if sheet_name not in entries_by_key[key]["sources"]:
                entries_by_key[key]["sources"][sheet_name] = row_idx

        if min_row != float('inf'):
            row_ranges[sheet_name] = {"min": min_row, "max": max_row}

    entries = list(entries_by_key.values())

    return {
        "metadata": {
            "sourceWorkbook": WORKBOOK_PATH.name,
            "includedSheets": list(TARGET_SHEETS.keys()),
            "sourceCounts": source_counts,
            "blankRows": blank_rows,
            "incompleteRows": incomplete_rows,
            "totalSourcePairs": sum(source_counts.values()),
            "uniqueEntries": len(entries),
            "rowRanges": row_ranges,
        },
        "entries": entries,
    }


def main() -> None:
    dataset = build_dataset()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(dataset, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(
        f"Wrote {dataset['metadata']['uniqueEntries']} entries to {OUTPUT_PATH.relative_to(ROOT)}"
    )


if __name__ == "__main__":
    main()
